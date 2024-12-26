#!/usr/bin/env python3
"""
Class count sheet maker.
"""
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def main():
    print("hello, world!")
    # Step 1: Load the Excel sheet with pandas
    df = pd.read_excel('input_file.xlsx')
    # Step 3: Create a new Excel workbook and add the data
    wb = Workbook()
    ws = wb.active

    # Add DataFrame to Excel (via openpyxl)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Step 4: Apply color formatting to cells based on the date column (assumed to be 'Date')
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        date_cell = row[0]  # Assuming the date is in the first column
        date_value = pd.to_datetime(date_cell.value)
        date_cell.fill = apply_color_based_on_date(date_value.date())

    # Step 5: Save the modified Excel file
    wb.save('output_file.xlsx')


# Step 2: Define a color fill based on date
def apply_color_based_on_date(date_value):
    today_date = pd.Timestamp.today().date()

    print(today_date)
    print(date_value)
    if date_value < today_date:  # For past dates
        return PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
    elif date_value == today_date:  # For today's date
        return PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Yellow
    else:  # For future dates
        return PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Green


if __name__ == "__main__":
    main()